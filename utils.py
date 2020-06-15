import io
import json
from typing import List

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.worksheet.worksheet import Worksheet


def load_from_file(filename: str) -> dict:
    with open(filename, 'r', encoding='UTF-8') as file:
        data = file.read()
        return json.loads(data)


def save_to_file(filename: str, data: dict) -> None:
    with open(filename, 'w+', encoding='UTF-8') as file:
        file.write(json.dumps(data, ensure_ascii=False))


class Table:

    def __init__(self, headers: list, rows: List[list]):
        self.headers = headers
        self.rows = rows


class ExcelBuilder:

    def __init__(self):
        self.tables: List[Table] = []
        self.images: List[Image] = []
        self.wb = Workbook()
        self.sheet: Worksheet = self.wb.active

    def add_table(self, headers: list, rows: List[list]):
        self.tables.append(Table(headers, rows))

    def add_image(self, image: io.BytesIO):
        self.images.append(Image(image))

    def _build_excel(self):
        current_row = 1
        for table in self.tables:
            for col, header in enumerate(table.headers):
                self.sheet.cell(current_row, col + 1, header)
            current_row += 1
            for row in table.rows:
                for col, value in enumerate(row):
                    self.sheet.cell(current_row, col + 1, value)
                current_row += 1
            current_row += 2

        for image in self.images:
            image.anchor = f'A{current_row}'
            self.sheet.add_image(image)

    def save(self, filename):
        self._build_excel()
        self.wb.save(filename)
